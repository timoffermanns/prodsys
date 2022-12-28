from typing import Dict, Union, Type, List, Optional
from pydantic import BaseModel, validator, Extra
import datetime

from prodsim import adapters

from prodsim.data_structures import (
    queue_data,
    resource_data,
    time_model_data,
    state_data,
    processes_data,
    material_data,
    sink_data,
    source_data,
)

import pandas as pd


class ProcessTime(BaseModel):
    machineDurationGroup: str
    taskDurationGroup: str
    duration: datetime.datetime

    @validator("duration", pre=True)
    def duration_to_timedelta(cls, v):
        return datetime.datetime.strptime(v[3:], "%H:%M:%S.%f")


class TransitionTime(BaseModel):
    jobTypeName: str
    transitionTimeGroup: str
    duration: datetime.datetime

    @validator("duration", pre=True)
    def duration_to_timedelta(cls, v: pd.Timestamp):
        if not v:
            return datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
        return v


class TaskType(BaseModel):
    name: str
    label: Optional[str]
    description: str
    requiredCapabilityNames: str
    setupGroup: str
    durationGroup: str

class Machine(BaseModel):
    name: str
    label: str
    description: str
    locationName: str
    shiftName: str
    setupGroup: str
    availableCapabilityNames: str
    setupGroup: Optional[str]
    durationGroup:str

    class Config:
            extra = "ignore"

class FlexisDataFrames(BaseModel):
    ApplicationName: pd.DataFrame
    Capability: pd.DataFrame
    Customer: pd.DataFrame
    Product: pd.DataFrame
    OrderType: pd.DataFrame
    JobType: pd.DataFrame
    Order: pd.DataFrame
    Location: pd.DataFrame
    Machine: pd.DataFrame
    TaskType: pd.DataFrame
    WorkPlan: pd.DataFrame
    SetupTime: pd.DataFrame
    ProcessTime: pd.DataFrame
    TransitionTime: pd.DataFrame
    CalendarParameter: pd.DataFrame
    DayPattern: pd.DataFrame
    Shift: pd.DataFrame
    NonworkingDay: pd.DataFrame
    Constraint: pd.DataFrame
    LinkType: pd.DataFrame
    Link: pd.DataFrame

    class Config:
        arbitrary_types_allowed = True


from openpyxl import load_workbook


def read_excel(file_path: str):
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames
    data = {}
    for sheet in sheets:
        data[sheet] = pd.DataFrame(
            wb[sheet].values, columns=[cell.value for cell in wb[sheet][1]],
        )
        data[sheet] = data[sheet].dropna(how="all")
        data[sheet] = data[sheet].iloc[1:]
        new_columns = [column for column in data[sheet].columns if column]
        data[sheet] = data[sheet][new_columns] 
    return data


class FlexisAdapter(adapters.Adapter):
    _capability_process_dict: Dict[str, List[str]] = {}

    def read_data(self, file_path: str):    
        input_data = read_excel(file_path)
        flexis_data_frames = FlexisDataFrames(**input_data)
        self.initialize_time_models(flexis_data_frames)
        self.initialize_process_models(flexis_data_frames)
        # TODO: add setup states here and breakdown states
        self.initialize_resource_models(flexis_data_frames)

    def get_object_from_data_frame(
        self, data_frame: pd.DataFrame, type: Type
    ) -> List[Type]:
        new_cols = [column for column in data_frame.columns if column]
        data_frame = data_frame[new_cols]
        data = data_frame.to_dict(orient="index")
        object_list = []
        for entry, entry_values in data.items():
            for key in entry_values:
                if ":" in key:
                    new_key = key.split(":")[0]
                    data[entry][new_key] = data[entry].pop(key)
            object_list.append(type(**data[entry]))

        return object_list

    def initialize_time_models(self, flexis_data_frames: FlexisDataFrames):
        process_time_data = self.get_object_from_data_frame(
            flexis_data_frames.ProcessTime, ProcessTime
        )
        for process_time in process_time_data:
            self.time_model_data.append(self.create_process_time_model(process_time))
        transition_time_data = self.get_object_from_data_frame(
            flexis_data_frames.TransitionTime, TransitionTime
        )
        for transition_time in transition_time_data:
            self.time_model_data.append(
                self.create_transport_time_model(transition_time)
            )

    def create_process_time_model(self, process_time: ProcessTime):
        return time_model_data.FunctionTimeModelData(
            ID=process_time.machineDurationGroup + "_" + process_time.taskDurationGroup,
            description=f"Process time of {process_time.machineDurationGroup} and {process_time.taskDurationGroup}",
            type=time_model_data.TimeModelEnum.FunctionTimeModel,
            distribution_function="constant",
            parameters=[
                (
                    process_time.duration
                    - datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
                ).total_seconds()
                / 60
            ],
            batch_size=100,
        )

    def create_transport_time_model(self, transition_time: TransitionTime):
        return time_model_data.FunctionTimeModelData(
            ID=transition_time.jobTypeName + "_" + transition_time.transitionTimeGroup,
            description=f"Process time of {transition_time.jobTypeName} and {transition_time.transitionTimeGroup}",
            type=time_model_data.TimeModelEnum.FunctionTimeModel,
            distribution_function="constant",
            parameters=[
                (
                    transition_time.duration
                    - datetime.datetime.strptime("00:00:00.000", "%H:%M:%S.%f")
                ).total_seconds()
                / 60
            ],
            batch_size=100,
        )

    def initialize_process_models(self, flexis_data_frames: FlexisDataFrames):
        task_type_data = self.get_object_from_data_frame(
            flexis_data_frames.TaskType, TaskType
        )
        machine_duration_groups = flexis_data_frames.Machine[
            "durationGroup:String"
        ].unique()
        for task_type in task_type_data:
            self.process_data += self.create_process_model(
                task_type, machine_duration_groups
            )

    def create_process_model(
        self, task_type: TaskType, machine_duration_groups: List[str]
    ):
        proceses_data_models = []
        for machine_duration_group in machine_duration_groups:
            time_model_ids = set([t.ID for t in self.time_model_data])
            time_model_id = machine_duration_group.split(":")[0] + "_" + task_type.durationGroup
            if time_model_id in time_model_ids:
                proceses_data_models.append(
                    processes_data.ProductionProcessData(
                        ID=task_type.name,
                        description=task_type.description,
                        type=processes_data.ProcessTypeEnum.ProductionProcesses,
                        time_model_id=time_model_id,
                    )
                )
            if not task_type.requiredCapabilityNames in self._capability_process_dict:
                self._capability_process_dict[task_type.requiredCapabilityNames] = []
            self._capability_process_dict[task_type.requiredCapabilityNames].append(time_model_id)
        return proceses_data_models
    
    def initialize_resource_models(self, flexis_data_frames: FlexisDataFrames):
        resource_data = self.get_object_from_data_frame(
            flexis_data_frames.Machine, Machine
        )
        for resource in resource_data:
            self.resource_data.append(self.create_resource_model(resource))

    def create_resource_model(self, resource: Machine):
        return resource_data.ProductionResourceData(
            ID=resource.name,
            description=resource.description,
            capacity=1,
            location=(0,0),
            controller="SimpleController",
            control_policy="FIFO",
            processes=self._capability_process_dict[resource.availableCapabilityNames],
            states=[]
        )

    def write_data(self, file_path: str):
        pass
